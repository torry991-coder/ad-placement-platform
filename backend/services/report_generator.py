"""Report generation service.

Generates multi-dimensional reports (daily/hourly summaries + time series)
for campaigns, with CSV and PDF export support.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Any
from uuid import uuid4

from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.performance import PerformanceMetric
from backend.models.campaign import Campaign

logger = logging.getLogger(__name__)

_AGG_METRICS = {"impressions", "clicks", "conversions", "spend", "revenue"}
_RATIO_METRICS = {"ctr", "cvr", "cpc", "cpa", "roas"}


# ── public API ──────────────────────────────────────────────────────────
async def generate_report(
    db: AsyncSession,
    report_type: str = "daily",
    campaign_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    metrics: list[str] | None = None,
) -> dict[str, Any]:
    """Generate a performance report for specified campaigns over a date range.

    Args:
        db: Async DB session.
        report_type: 'daily' or 'hourly'.
        campaign_id: Single campaign ID or None for all campaigns.
        date_from: Start date YYYY-MM-DD (defaults to 7 days ago).
        date_to: End date YYYY-MM-DD (defaults to today).
        metrics: List of metric names to include. Defaults to all.

    Returns:
        Dict with: report_id, report_type, generated_at, date_range,
        granularity, campaigns, summary (totals), time_series, metrics_included.
    """
    if not metrics:
        metrics = ["impressions", "clicks", "conversions", "spend", "revenue",
                    "ctr", "cvr", "cpc", "cpa", "roas"]

    granularity = "hourly" if report_type == "hourly" else "daily"

    # Default date range: last 7 days
    now = datetime.now(timezone.utc)
    if not date_to:
        date_to = now.strftime("%Y-%m-%d")
    if not date_from:
        from datetime import timedelta
        date_from = (now - timedelta(days=7)).strftime("%Y-%m-%d")

    # Build campaign ID list
    campaign_ids: list[int] = []
    if campaign_id:
        campaign_ids = [campaign_id]

    # Resolve campaign names
    campaign_names: dict[int, str] = {}
    if campaign_ids:
        result = await db.execute(
            select(Campaign.id, Campaign.name).where(Campaign.id.in_(campaign_ids))
        )
        for row in result.all():
            campaign_names[row[0]] = row[1]
    else:
        # All campaigns
        result = await db.execute(select(Campaign.id, Campaign.name))
        for row in result.all():
            campaign_ids.append(row[0])
            campaign_names[row[0]] = row[1]

    report_id = str(uuid4())[:8]

    # ── Fetch performance data ──────────────────────────────────────────
    query = select(PerformanceMetric).where(
        and_(
            PerformanceMetric.date >= date_from,
            PerformanceMetric.date <= date_to,
        )
    )
    if campaign_ids:
        query = query.where(PerformanceMetric.campaign_id.in_(campaign_ids))
    if granularity == "daily":
        query = query.where(PerformanceMetric.hour.is_(None))

    result = await db.execute(
        query.order_by(PerformanceMetric.date.asc(), PerformanceMetric.hour.asc())
    )
    rows = result.scalars().all()

    # ── Build summary ───────────────────────────────────────────────────
    totals = {
        "impressions": 0, "clicks": 0, "conversions": 0,
        "spend": 0.0, "revenue": 0.0,
    }
    campaign_summaries: dict[int, dict[str, Any]] = {}
    time_series: list[dict[str, Any]] = []

    for r in rows:
        cid = r.campaign_id
        totals["impressions"] += r.impressions or 0
        totals["clicks"] += r.clicks or 0
        totals["conversions"] += r.conversions or 0
        totals["spend"] += float(r.spend or 0)
        totals["revenue"] += float(r.revenue or 0)

        if cid not in campaign_summaries:
            campaign_summaries[cid] = {
                "campaign_id": cid,
                "campaign_name": campaign_names.get(cid, f"Campaign #{cid}"),
                "impressions": 0, "clicks": 0, "conversions": 0,
                "spend": 0.0, "revenue": 0.0,
            }
        cs = campaign_summaries[cid]
        cs["impressions"] += r.impressions or 0
        cs["clicks"] += r.clicks or 0
        cs["conversions"] += r.conversions or 0
        cs["spend"] += float(r.spend or 0)
        cs["revenue"] += float(r.revenue or 0)

        # Time series point
        time_series.append({
            "date": r.date,
            "hour": r.hour,
            "campaign_id": cid,
            "campaign_name": campaign_names.get(cid, ""),
            "platform": r.platform,
            "impressions": r.impressions or 0,
            "clicks": r.clicks or 0,
            "conversions": r.conversions or 0,
            "spend": round(float(r.spend or 0), 2),
            "revenue": round(float(r.revenue or 0), 2),
            "ctr": round(r.ctr or 0, 4),
            "cvr": round(r.cvr or 0, 4),
            "cpc": round(r.cpc or 0, 4),
            "cpa": round(r.cpa or 0, 4),
            "roas": round(r.roas or 0, 4),
        })

    # Compute derived metrics per campaign
    for cs in campaign_summaries.values():
        imp = cs["impressions"]
        clk = cs["clicks"]
        conv = cs["conversions"]
        sp = cs["spend"]
        rev = cs["revenue"]
        cs["ctr"] = round(clk / imp * 100, 4) if imp > 0 else 0.0
        cs["cvr"] = round(conv / clk * 100, 4) if clk > 0 else 0.0
        cs["cpc"] = round(sp / clk, 4) if clk > 0 else 0.0
        cs["cpa"] = round(sp / conv, 4) if conv > 0 else 0.0
        cs["roas"] = round(rev / sp, 4) if sp > 0 else 0.0

    summary = _build_summary(totals, metrics)

    return {
        "report_id": report_id,
        "report_type": report_type,
        "generated_at": now,
        "date_range": {"from": date_from, "to": date_to},
        "granularity": granularity,
        "campaigns": list(campaign_summaries.values()),
        "summary": summary,
        "time_series": time_series,
        "total_periods": len(time_series),
        "metrics_included": metrics,
    }


# ── Export functions (return string content, no file path needed) ──────
def _export_csv_bytes(data: dict[str, Any]) -> bytes:
    """Export time_series data to CSV bytes (fallback when openpyxl missing)."""
    time_series = data.get("time_series", [])
    if not time_series:
        return b"No data available"
    import csv as _csv, io as _io
    output = _io.StringIO()
    fieldnames = list(time_series[0].keys())
    writer = _csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(time_series)
    return output.getvalue().encode("utf-8")


def export_excel(data: dict[str, Any]) -> bytes:
    """Export report to a multi-sheet Excel workbook (.xlsx bytes).

    Sheets:
      - 概览 (Summary): KPI overview
      - 活动明细 (Campaigns): Per-campaign breakdown
      - 时间序列 (Time Series): Daily/hourly trend data
      - 平台分布 (Platforms): Platform-level breakdown

    Returns: Excel file as bytes (ready for StreamingResponse).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV bytes if openpyxl not installed
        return _export_csv_bytes(data)

    wb = openpyxl.Workbook()

    # ── Styles ──────────────────────────────────────────────────────────
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    def style_header(ws, headers: list[str], row: int = 1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    # ── Sheet 1: 概览 ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "概览"
    summary = data.get("summary", {})
    headers = ["指标", "数值"]
    style_header(ws1, headers)
    rows = [
        ["展示量", summary.get("impressions", 0)],
        ["点击量", summary.get("clicks", 0)],
        ["转化量", summary.get("conversions", 0)],
        ["花费 (¥)", round(float(summary.get("spend", 0)), 2)],
        ["收入 (¥)", round(float(summary.get("revenue", 0)), 2)],
        ["CTR", f"{summary.get('ctr', 0):.2f}%"],
        ["CVR", f"{summary.get('cvr', 0):.2f}%"],
        ["CPC (¥)", round(float(summary.get("cpc", 0)), 4)],
        ["CPA (¥)", round(float(summary.get("cpa", 0)), 4)],
        ["ROAS", f"{summary.get('roas', 0):.2f}x"],
    ]
    for r, (label, value) in enumerate(rows, 2):
        ws1.cell(row=r, column=1, value=label).font = cell_font
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2, value=value).font = cell_font
        ws1.cell(row=r, column=2).alignment = cell_align
        ws1.cell(row=r, column=2).border = thin_border
    auto_width(ws1)

    # ── Sheet 2: 活动明细 ──────────────────────────────────────────────
    ws2 = wb.create_sheet("活动明细")
    campaigns = data.get("campaigns", [])
    if campaigns:
        headers2 = ["活动名称", "展示量", "点击量", "转化量", "花费", "收入", "CTR", "CVR", "ROAS"]
        style_header(ws2, headers2)
        for r, c in enumerate(campaigns, 2):
            values = [
                c.get("campaign_name", ""),
                int(c.get("impressions", 0)),
                int(c.get("clicks", 0)),
                int(c.get("conversions", 0)),
                round(float(c.get("spend", 0)), 2),
                round(float(c.get("revenue", 0)), 2),
                f"{c.get('ctr', 0):.2f}%",
                f"{c.get('cvr', 0):.2f}%",
                f"{c.get('roas', 0):.2f}x",
            ]
            for col, v in enumerate(values, 1):
                cell = ws2.cell(row=r, column=col, value=v)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border
    auto_width(ws2)

    # ── Sheet 3: 时间序列 ──────────────────────────────────────────────
    ws3 = wb.create_sheet("时间序列")
    ts = data.get("time_series", [])
    if ts:
        headers3 = ["日期", "小时", "展示量", "点击量", "转化量", "花费", "收入", "CTR", "CVR", "ROAS"]
        style_header(ws3, headers3)
        for r, row_data in enumerate(ts, 2):
            values = [
                row_data.get("date", ""),
                row_data.get("hour", ""),
                int(row_data.get("impressions", 0)),
                int(row_data.get("clicks", 0)),
                int(row_data.get("conversions", 0)),
                round(float(row_data.get("spend", 0)), 2),
                round(float(row_data.get("revenue", 0)), 2),
                f"{row_data.get('ctr', 0):.2f}%",
                f"{row_data.get('cvr', 0):.2f}%",
                f"{row_data.get('roas', 0):.2f}x",
            ]
            for col, v in enumerate(values, 1):
                cell = ws3.cell(row=r, column=col, value=v)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border
    auto_width(ws3)

    # ── Save to bytes ──────────────────────────────────────────────────
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
    """Export report time_series data to CSV string.

    Returns: CSV content as string.
    """
    time_series = data.get("time_series", [])
    if not time_series:
        return "No data available"

    output = io.StringIO()
    fieldnames = list(time_series[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(time_series)
    return output.getvalue()


def export_pdf(data: dict[str, Any]) -> str:
    """Generate a simple HTML-based PDF-like export.

    In production, use WeasyPrint for real PDF generation.
    Returns HTML content that browsers can render as PDF.
    """
    summary = data.get("summary", {})
    date_range = data.get("date_range", {})
    campaigns = data.get("campaigns", [])

    html_parts = [
        "<html><head><meta charset='utf-8'>",
        "<style>body{font-family:Arial,sans-serif;margin:40px}",
        "h1{color:#1a1a2e}h2{color:#16213e}",
        "table{border-collapse:collapse;width:100%;margin:20px 0}",
        "th,td{border:1px solid #ddd;padding:8px;text-align:right}",
        "th{background:#1a1a2e;color:white}",
        "</style></head><body>",
        f"<h1>Ad Placement Report</h1>",
        f"<p><strong>Date Range:</strong> {date_range.get('from','')} → {date_range.get('to','')}</p>",
        f"<p><strong>Generated:</strong> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC</p>",
        "<h2>Summary</h2><table>",
        "<tr><th>Metric</th><th>Value</th></tr>",
    ]
    for k, v in summary.items():
        html_parts.append(f"<tr><td>{k}</td><td>{v}</td></tr>")
    html_parts.append("</table>")

    if campaigns:
        html_parts.append("<h2>Campaign Breakdown</h2><table>")
        keys = ["campaign_name", "impressions", "clicks", "conversions", "spend", "revenue", "ctr", "cvr", "roas"]
        html_parts.append("<tr>" + "".join(f"<th>{k}</th>" for k in keys) + "</tr>")
        for c in campaigns:
            html_parts.append("<tr>" + "".join(f"<td>{c.get(k, '')}</td>" for k in keys) + "</tr>")
        html_parts.append("</table>")

    html_parts.append("</body></html>")
    return "\n".join(html_parts)


# ── helpers ─────────────────────────────────────────────────────────────
def _empty_summary() -> dict[str, Any]:
    return {
        "impressions": 0, "clicks": 0, "conversions": 0,
        "spend": 0.0, "revenue": 0.0,
        "ctr": 0.0, "cvr": 0.0, "cpc": 0.0, "cpa": 0.0, "roas": 0.0,
    }


def _build_summary(totals: dict[str, float], metrics: list[str]) -> dict[str, Any]:
    summary: dict[str, Any] = {}
    impressions = totals.get("impressions", 0)
    clicks = totals.get("clicks", 0)
    conversions = totals.get("conversions", 0)
    spend = totals.get("spend", 0)
    revenue = totals.get("revenue", 0)

    if "impressions" in metrics:
        summary["impressions"] = int(impressions)
    if "clicks" in metrics:
        summary["clicks"] = int(clicks)
    if "conversions" in metrics:
        summary["conversions"] = int(conversions)
    if "spend" in metrics:
        summary["spend"] = round(spend, 2)
    if "revenue" in metrics:
        summary["revenue"] = round(revenue, 2)
    if "ctr" in metrics:
        summary["ctr"] = round(clicks / impressions * 100.0, 4) if impressions > 0 else 0.0
    if "cvr" in metrics:
        summary["cvr"] = round(conversions / clicks * 100.0, 4) if clicks > 0 else 0.0
    if "cpc" in metrics:
        summary["cpc"] = round(spend / clicks, 4) if clicks > 0 else 0.0
    if "cpa" in metrics:
        summary["cpa"] = round(spend / conversions, 4) if conversions > 0 else 0.0
    if "roas" in metrics:
        summary["roas"] = round(revenue / spend, 4) if spend > 0 else 0.0

    return summary
